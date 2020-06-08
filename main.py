import os.path
from os import walk
#pip install openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
#pip instal pillow
import PIL
from PIL import Image as Pimage

#função para pegar dados destinada a imagem
def Informa(texto):
    #função retorna Base, altura e nome da imagem, nesta ordem
    dim = texto[1:texto.find(')')]
    nome = texto[texto.find(')') + 1:len(texto)]
    base = dim[0:dim.find('x')]
    altura = dim[dim.find('x') + 1:len(dim)]
    try:
        base = int(base)
        altura = int(altura)
        return base, altura, nome
    except:
        print('Não é possivel ler o tamanho da imagem desejada')
        return nome
        pass

print('##########################################################################################')
print('#######  ESTE PROGRAMA REALIZA O PREENCHIMENTO AUTOMATICO DE IMAGENS EM DOCUMENTOS #######')
print('##########################################################################################')
print('######################################################### CREATE BY: BRUNO B. LINHARES ###')
print('##########################################################################################')
print('\n')

# endereço do arquivo py print(os.path.realpath(__file__))
# Pasta lacal do arquivo executado print(os.path.abspath(""))
lista_ext_compati = ['.xlsx','.xlsm','.docx']
minhapasta = os.path.abspath("")
#lista os arquivos existentes dentro da pasta atual sem pastas e subpastas
for (dirpath, dirnames, lista_de_arquivos) in walk(minhapasta):
    break
#verifica quais arquivos listados são compativeis
lista_de_arquivos_compat = []
for arquivo in lista_de_arquivos:
    for ext in lista_ext_compati:
        if arquivo.endswith(ext) == True :
            lista_de_arquivos_compat.append(arquivo)
#verificar se existe arquivos compativeis
#se não houver
if len(lista_de_arquivos_compat) == 0:
    print('NÃO HÁ ARQUIVOS COMPATIVEIS ...\n')
    print('NÃO É POSSIVEL CONTINUAR A EXECUÇÃO DO PROGRAMA ...\n')
else:
    #se houver, o programa continua
    print('Foram encontrados ' + str(len(lista_de_arquivos_compat)) + ' arquivos compativeis...')
    print ('Selecione o documento que deseja preencher \n')
    #Lista os arquivos para escolha do usuário
    i = 1
    for iten in lista_de_arquivos_compat:
        print(str(i) + ' - ' + iten +':')
        i = i + 1
    opcao = False
    #converte a opção em inteiro
    n_opcao = input(u'Digite a opção desejada:')
    while opcao == False:
        try :
            n_opcao = int(n_opcao)
            if n_opcao < (i+1):
                print('\nOpção válida!')
                print('O arquivo ' + lista_de_arquivos_compat[n_opcao-1] + ' foi selecionado.')
                opcao = True
            else:
                n_opcao = input('\nOpção invalida! Digite novamente:')
        except:
            n_opcao = input('\nOpção invalida! Digite novamente:')
        #nome do arquivo alvo
        end_arq_alvo = os.path.abspath("") +'\\' + lista_de_arquivos_compat[n_opcao-1]
        print('\nO arquivo localizado em: ' + end_arq_alvo + ' será editado...')

    #carregar aquivo a ser editado
    #carregar se excel
    try:
        carr_doc = load_workbook(filename=end_arq_alvo, data_only=True)
        lista_de_abas_disp = carr_doc.sheetnames
        print('Foram encontrados ' + str(len(lista_de_abas_disp)) + ' abas na planilha...')
        print('Selecione a aba que deseja preencher \n')
        # Lista de abas para escolha do usuário
        i = 1
        for iten in lista_de_abas_disp:
            print(str(i) + ' - ' + iten + ':')
            i = i + 1
        #ler opção
        opcao = False
        n_opcao = input(u'Digite a opção desejada:')
        while opcao == False:
            try:
                n_opcao = int(n_opcao)
                if n_opcao < (i + 1):
                    print('\nOpção válida!')
                    print('A aba ' + lista_de_abas_disp[n_opcao - 1] + ' foi selecionado.')
                    opcao = True
                else:
                    n_opcao = input('\nOpção invalida! Digite novamente:')
            except:
                n_opcao = input('\nOpção invalida! Digite novamente:')
        #carregar a aba selecionada
        print('Carregando aba selecionanda...')
        carr_doc_aba = carr_doc[lista_de_abas_disp[n_opcao - 1]]
        #lista de arquivos a serem apagados posteriormente
        lixeira = []
        print('Percorrendo celulas a procura de imagens...\n')
        #percorrer todas as colunas da planilha
        for coluna in range(1,carr_doc_aba.max_column):
            #percorrer linhas
            for linha in range(1,carr_doc_aba.max_row):
                nome_da_celula = str(get_column_letter(coluna) + str(linha))
                valor_da_celula = str(carr_doc_aba[nome_da_celula].value)
                if valor_da_celula.endswith('.jpg') == True or valor_da_celula.endswith('.png') == True :
                    print('Extenção de imagem encontrada na celula '+ nome_da_celula + '.')
                    data = Informa(valor_da_celula)

                    if type(data) == str:
                        base_da_imagem_thumbnail = 255
                        altura_da_imagem_thumbnail = 255
                        nome_da_imagem = data
                    if type(data) == tuple:
                        base_da_imagem_thumbnail = data[0]
                        altura_da_imagem_thumbnail = data[1]
                        nome_da_imagem = data[2]


                    end_da_img = os.path.abspath("")+'\\img\\'+nome_da_imagem
                    print('Carregando imagem de : '+end_da_img)
                    try:
                        imagem = Pimage.open(end_da_img)
                        print('Imagen Carregada com Sucesso!')
                        '''
                        #recortando e redimensionando imagem
                        #marca dágua pode ser removida
                        base_da_imagem = int(imagem.size[0])
                        altura_da_imagem = int(imagem.size[1])
                        #corta imagem centralizada
                        # altura nova (relação base/altura
                        n_altura = int(base_da_imagem / 1.78)
                        # cortar imagem centralizada
                        c_dim = int((altura_da_imagem - n_altura) / 2)
                        imagem_recort = imagem.crop((0, c_dim, base_da_imagem, n_altura + c_dim))
                        imagem = imagem_recort
                        '''
                        imagem.thumbnail((base_da_imagem_thumbnail, altura_da_imagem_thumbnail))
                        imagem.save(nome_da_imagem)
                        lixeira.append(nome_da_imagem)
                        imagem = Image(nome_da_imagem)
                        carr_doc_aba.add_image(imagem, nome_da_celula)
                        print('Imagem Inserida com Sucesso!\n')

                    except:
                        print('Não foi possível carregar a imagem!')


        print('Salvando a Planilha em: ' + end_arq_alvo + '...')
        carr_doc.save(end_arq_alvo)
        carr_doc.close()
        print('Arquivo Salvo com Sucesso!\n')
        print('Apagando arquivos temporarios...')
        for lixo in lixeira:
            try:
                os.remove(lixo)
            except:
                pass

        print('Arquivos temporarios removidos com sucesso...\n')
        print('##########################################################################################')
        print('################################## ROTINA  CONCLUIDA ! ##################################')
        print('##########################################################################################')
    except:

        print('ARQUIVOS EM MICRISOFT WORD AINDA NÃO FORAM IMPLEMENTADOS...')
        pass


